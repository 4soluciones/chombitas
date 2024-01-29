import datetime
import decimal

from django.db.models import F, Prefetch, Subquery, OuterRef, Sum
from django.db.models.functions import Coalesce
from django.shortcuts import render
from django.http.response import HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.accounting.models import CashFlow
from apps.buys.models import Purchase, PurchaseDetail, Requirement_buys, RequirementDetail_buys
from apps.hrm.models import Subsidiary
from apps.hrm.views import get_subsidiary_by_user, User
from apps.sales.models import Product, OrderDetail, Order, Supplier, SubsidiaryStore
from apps.sales.views import excel_order_by_units


class StockMin(TemplateView):

    def get(self, request, *args, **kwargs):
        query = Product.objects.filter(stock__lte=F('minimum'))
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = 'STOCK DE PRODUCTOS' + str(cont)  # hoja
            bandera = False
        else:
            ws = wb.create_sheet('STOCK DE PRODUCTOS' + str(cont))
        # Crear el título en la hoja
        my_date = datetime.datetime.now()
        date_now = my_date.strftime("%d-%m-%Y")
        hour_now = my_date.strftime("%H:%M:%S")
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = 'PRODUCTOS CON STOCK AL MINIMO ' + ' FECHA: ' + str(date_now) + ' HORA: ' + str(hour_now)

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:N1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 9
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 25

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'CODIGO'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'NOMBRE PRODUCTO'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'DESCRIPCION PRODUCTO'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'MARCA'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'FAMILIA'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'TIPO'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'ANCHO'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'LARGO'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'ALTO'

        ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K3'] = 'ALMACEN'

        ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L3'] = 'STOCK'

        ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M3'] = 'MINIMO'

        ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['N3'] = 'PRODUCTO RELACIONADO'

        for p in query.order_by('id'):
            # Pintamos los datos en el reporte
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            stock_cell = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")
            text_color = Font(name='Calibri', size=9)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="center")
            ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).border = border_cell
            ws.cell(row=row, column=2).fill = color_cell
            ws.cell(row=row, column=2).font = text_color
            ws.cell(row=row, column=2).value = p.code

            ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=3).border = border_cell
            ws.cell(row=row, column=3).fill = color_cell
            ws.cell(row=row, column=3).font = text_color
            ws.cell(row=row, column=3).value = p.name

            ws.cell(row=row, column=4).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=4).border = border_cell
            ws.cell(row=row, column=4).fill = color_cell
            ws.cell(row=row, column=4).font = text_color
            ws.cell(row=row, column=4).value = p.description

            ws.cell(row=row, column=5).alignment = align_cell
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = p.brand.name

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = p.family.name

            ws.cell(row=row, column=7).alignment = align_cell
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = p.get_type_display()

            ws.cell(row=row, column=8).alignment = align_cell
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = p.width

            ws.cell(row=row, column=9).alignment = align_cell
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = p.length

            ws.cell(row=row, column=10).alignment = align_cell
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = p.height

            ws.cell(row=row, column=11).alignment = align_cell
            ws.cell(row=row, column=11).border = border_cell
            ws.cell(row=row, column=11).fill = color_cell
            ws.cell(row=row, column=11).font = text_color
            ws.cell(row=row, column=11).value = p.store

            ws.cell(row=row, column=12).alignment = align_cell
            ws.cell(row=row, column=12).border = border_cell
            ws.cell(row=row, column=12).fill = stock_cell
            ws.cell(row=row, column=12).font = text_color
            ws.cell(row=row, column=12).value = p.stock

            ws.cell(row=row, column=13).alignment = align_cell
            ws.cell(row=row, column=13).border = border_cell
            ws.cell(row=row, column=13).fill = color_cell
            ws.cell(row=row, column=13).font = text_color
            ws.cell(row=row, column=13).value = p.minimum

            ws.cell(row=row, column=14).alignment = align_cell
            ws.cell(row=row, column=14).border = border_cell
            ws.cell(row=row, column=14).fill = color_cell
            ws.cell(row=row, column=14).font = text_color
            ws.cell(row=row, column=14).value = p.relation

            cont += 1
            row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "MinStock.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def report_sales_excel(request, init=None, end=None, pk=0):
    if init and end:
        start_date = datetime.datetime.strptime(init, "%Y-%m-%d")
        formatted_init = start_date.strftime("%d-%m-%Y")
        end_date = datetime.datetime.strptime(end, "%Y-%m-%d")
        formatted_end = end_date.strftime("%d-%m-%Y")
        wb = Workbook()
        bandera = True
        cont = 1
        if bandera:
            ws = wb.active
            ws.title = str("REPORTE DE VENTAS")  # hoja
            bandera = False
        else:
            ws = wb.create_sheet(str("REPORTE DE VENTAS") + str(cont))
        # Crear el título en la hoja
        row = 2
        if pk == '0' or pk == 0:
            subsidiary_set = Subsidiary.objects.all()
            ws['B{}'.format(row)].alignment = Alignment(horizontal="center", vertical="center")
            ws['B{}'.format(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B{}'.format(row)].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['B{}'.format(row)].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
            ws['B{}'.format(row)] = str("REPORTE DE VENTAS DEL ") + str(formatted_init) + " AL " + str(formatted_end)
        else:
            subsidiary_set = Subsidiary.objects.filter(id=int(pk))

            ws['B{}'.format(row)].alignment = Alignment(horizontal="center", vertical="center")
            ws['B{}'.format(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B{}'.format(row)].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['B{}'.format(row)].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
            ws['B{}'.format(row)] = str("REPORTE DE VENTAS DEL ") + str(formatted_init) + " AL " + str(
                formatted_end) + " DE LA SEDE DE " + str(subsidiary_set.first().name)
        # Cambiar caracteristicas de las celdas
    ws.merge_cells('B{}:E{}'.format(row, row))

    ws.row_dimensions[row].height = 35

    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25

    # Crear la cabecera
    row += 1
    ws['B{}'.format(row)].alignment = Alignment(horizontal="center", vertical="center")
    ws['B{}'.format(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                          top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B{}'.format(row)].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['B{}'.format(row)].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['B{}'.format(row)] = 'Nº'

    ws['C{}'.format(row)].alignment = Alignment(horizontal="center", vertical="center")
    ws['C{}'.format(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                          top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C{}'.format(row)].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['C{}'.format(row)].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['C{}'.format(row)] = 'SEDE'

    ws['D{}'.format(row)].alignment = Alignment(horizontal="center", vertical="center")
    ws['D{}'.format(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                          top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D{}'.format(row)].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['D{}'.format(row)].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['D{}'.format(row)] = 'TOTAL VENTA'

    ws['E{}'.format(row)].alignment = Alignment(horizontal="center", vertical="center")
    ws['E{}'.format(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                          top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['E{}'.format(row)].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['E{}'.format(row)].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['E{}'.format(row)] = 'TOTAL COBRANZA'
    ws.row_dimensions[row].height = 30
    row += 1
    if pk == '0' or pk == 0:
        subsidiary_set = Subsidiary.objects.all()
    else:
        subsidiary_set = Subsidiary.objects.filter(id=int(pk))
    t_sales = 0
    t_cash = 0
    for i, s in enumerate(subsidiary_set):
        t = Order.objects.filter(
            subsidiary_store__subsidiary_id=s.id,
            create_at__date__range=(init, end), type__in=['V', 'R']
        ).aggregate(r=Coalesce(Sum('total'), decimal.Decimal(0.00)))

        c = CashFlow.objects.filter(
            cash__subsidiary_id=s.id,
            type='E',
            transaction_date__range=(init, end)
        ).aggregate(r=Coalesce(Sum('total'), decimal.Decimal(0.00)))
        color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        text_color = Font(name='Calibri', size=12, bold=True)
        border_cell = Border(left=Side(border_style="thin"),
                             right=Side(border_style="thin"),
                             top=Side(border_style="thin"),
                             bottom=Side(border_style="thin"))
        align_cell = Alignment(horizontal="center")
        ws.cell(row=row, column=2).alignment = align_cell
        ws.cell(row=row, column=2).border = border_cell
        ws.cell(row=row, column=2).fill = color_cell
        ws.cell(row=row, column=2).font = text_color
        ws.cell(row=row, column=2).value = i + 1

        color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        text_color = Font(name='Calibri', size=12, bold=True)
        border_cell = Border(left=Side(border_style="thin"),
                             right=Side(border_style="thin"),
                             top=Side(border_style="thin"),
                             bottom=Side(border_style="thin"))
        align_cell = Alignment(horizontal="left")
        ws.cell(row=row, column=3).alignment = align_cell
        ws.cell(row=row, column=3).border = border_cell
        ws.cell(row=row, column=3).fill = color_cell
        ws.cell(row=row, column=3).font = text_color
        ws.cell(row=row, column=3).value = s.name

        color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        text_color = Font(name='Calibri', size=12, bold=True)
        border_cell = Border(left=Side(border_style="thin"),
                             right=Side(border_style="thin"),
                             top=Side(border_style="thin"),
                             bottom=Side(border_style="thin"))
        align_cell = Alignment(horizontal="right")
        ws.cell(row=row, column=4).alignment = align_cell
        ws.cell(row=row, column=4).border = border_cell
        ws.cell(row=row, column=4).fill = color_cell
        ws.cell(row=row, column=4).font = text_color
        ws.cell(row=row, column=4).value = "{:.2f}".format(round(float(t['r']), 2))

        color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        text_color = Font(name='Calibri', size=12, bold=True)
        border_cell = Border(left=Side(border_style="thin"),
                             right=Side(border_style="thin"),
                             top=Side(border_style="thin"),
                             bottom=Side(border_style="thin"))
        align_cell = Alignment(horizontal="right")
        ws.cell(row=row, column=5).alignment = align_cell
        ws.cell(row=row, column=5).border = border_cell
        ws.cell(row=row, column=5).fill = color_cell
        ws.cell(row=row, column=5).font = text_color
        ws.cell(row=row, column=5).value = "{:.2f}".format(round(float(c['r']), 2))
        row += 1
        t_sales += float(t['r'])
        t_cash += float(c['r'])
    color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
    text_color = Font(name='Calibri', size=13, bold=True)
    border_cell = Border(left=Side(border_style="thin"),
                         right=Side(border_style="thin"),
                         top=Side(border_style="thin"),
                         bottom=Side(border_style="thin"))
    align_cell = Alignment(horizontal="right")
    ws.cell(row=row, column=2).alignment = align_cell
    ws.cell(row=row, column=2).border = border_cell
    ws.cell(row=row, column=2).fill = color_cell
    ws.cell(row=row, column=2).font = text_color
    ws.cell(row=row, column=2).value = "TOTAL"
    ws.merge_cells('B{}:C{}'.format(row, row))

    ws.cell(row=row, column=4).alignment = align_cell
    ws.cell(row=row, column=4).border = border_cell
    ws.cell(row=row, column=4).fill = color_cell
    ws.cell(row=row, column=4).font = text_color
    ws.cell(row=row, column=4).value = "{:.2f}".format(round(float(t_sales), 2))

    ws.cell(row=row, column=5).alignment = align_cell
    ws.cell(row=row, column=5).border = border_cell
    ws.cell(row=row, column=5).fill = color_cell
    ws.cell(row=row, column=5).font = text_color
    ws.cell(row=row, column=5).value = "{:.2f}".format(round(float(t_cash), 2))

    nombre_archivo = "{} - {}.xlsx".format("REPORTE DE VENTAS Y COBRANZAS", str(1))
    # Definir el tipo de respuesta que se va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


def report_category_product(request, year=None):
    if year:
        month_names = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SETIEMBRE', 'OCTUBRE',
                       'NOVIEMBRE', 'DICIEMBRE']
        purchase_dict = []
        sum_float_purchases_sum_total = 0
        sector = Supplier._meta.get_field('sector').choices
        sum_month = [0] * len(month_names)
        total_total = 0
        wb = Workbook()
        bandera = True
        cont = 1
        if bandera:
            ws = wb.active
            ws.title = str("REPORTE DE COMPRAS")  # hoja
            bandera = False
        else:
            ws = wb.create_sheet(str("REPORTE") + str(cont))
        # Crear el título en la hoja
        my_date = datetime.datetime.now()
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B2'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B2'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B2'] = str("REPORTE DE COMPRAS POR RUBRO Y MES DEL AÑO ") + str(year)

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B2:O2')

        ws.row_dimensions[2].height = 35

        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15

        # Crear la cabecera
        ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B4'] = 'RUBRO'

        ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C4'] = 'ENERO'

        ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D4'] = 'FEBRERO'

        ws['E4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E4'] = 'MARZO'

        ws['F4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F4'] = 'ABRIL'

        ws['G4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G4'] = 'MAYO'

        ws['H4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H4'] = 'JUNIO'

        ws['I4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I4'] = 'JULIO'

        ws['J4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J4'] = 'AGOSTO'

        ws['K4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K4'] = 'SETIEMBRE'

        ws['L4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L4'] = 'OCTUBRE'

        ws['M4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M4'] = 'NOVIEMBRE'

        ws['N4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['N4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['N4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['N4'] = 'DICIEMBRE'

        ws['O4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['O4'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['O4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['O4'] = 'TOTAL AÑOS'

        ws.row_dimensions[4].height = 30

        counter_for = 0
        row = 5
        for c in sector:
            value, label = c
            sum_total_year = 0
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            text_color = Font(name='Calibri', size=12, bold=True)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="left")
            ws.cell(row=row + sector.index(c), column=2).alignment = align_cell
            ws.cell(row=row + sector.index(c), column=2).border = border_cell
            ws.cell(row=row + sector.index(c), column=2).fill = color_cell
            ws.cell(row=row + sector.index(c), column=2).font = text_color
            ws.cell(row=row + sector.index(c), column=2).value = label
            for m in month_names:
                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=12)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="right")
                if value == 'G':
                    requirement_set = Requirement_buys.objects.filter(status='2', type='M', status_pay='2',
                                                                      approval_date__year=year,
                                                                      approval_date__month=month_names.index(
                                                                          m) + 1).annotate(
                        sum_total=Subquery(
                            RequirementDetail_buys.objects.filter(requirement_buys_id=OuterRef('id')).annotate(
                                r=Sum(F('quantity') * F('price_pen'))).values('r')[:1])).aggregate(Sum('sum_total'))
                    requirement_sum_total = requirement_set['sum_total__sum']
                    if requirement_sum_total is not None:
                        float_requirement_sum_total = float(requirement_sum_total)
                    else:
                        float_requirement_sum_total = float(0)
                    sum_total_year += float_requirement_sum_total
                    sum_month[month_names.index(m)] += float_requirement_sum_total
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).alignment = align_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).border = border_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).fill = color_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).font = text_color
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).value = "{:.2f}".format(round(
                        decimal.Decimal(float_requirement_sum_total), 2))
                elif value == 'PP':
                    salary_total = CashFlow.objects.filter(salary__year=year,
                                                           salary__month=month_names.index(m) + 1).aggregate(
                        r=Coalesce(Sum('total'), decimal.Decimal(0.00))).get('r')
                    if salary_total is not None:
                        float_salary_total = float(salary_total)
                    else:
                        float_salary_total = float(0)
                    sum_total_year += float_salary_total
                    sum_month[month_names.index(m)] += float_salary_total
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).alignment = align_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).border = border_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).fill = color_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).font = text_color
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).value = "{:.2f}".format(round(
                        decimal.Decimal(float_salary_total), 2))
                else:
                    purchase_set = Purchase.objects.filter(
                        purchase_date__month=month_names.index(m) + 1, purchase_date__year=year, supplier__sector=value,
                        status='A'
                    ).select_related('supplier').annotate(
                        sum_total=Subquery(
                            PurchaseDetail.objects.filter(purchase_id=OuterRef('id')).annotate(
                                return_sum_total=Sum(F('quantity') * F('price_unit'))).values('return_sum_total')[:1]
                        )
                    ).aggregate(Sum('sum_total'))

                    purchases_sum_total = purchase_set['sum_total__sum']

                    if purchases_sum_total is not None:
                        float_purchases_sum_total = float(purchases_sum_total)
                    else:
                        float_purchases_sum_total = 0
                    sum_total_year += float_purchases_sum_total
                    sum_month[month_names.index(m)] += float_purchases_sum_total
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).alignment = align_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).border = border_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).fill = color_cell
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).font = text_color
                    ws.cell(row=row + sector.index(c), column=month_names.index(m) + 3).value = "{:.2f}".format(round(
                        decimal.Decimal(float_purchases_sum_total), 2))
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            text_color = Font(name='Calibri', size=12, bold=True)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="right")
            ws.cell(row=row + sector.index(c), column=month_names.index(m) + 4).alignment = align_cell
            ws.cell(row=row + sector.index(c), column=month_names.index(m) + 4).border = border_cell
            ws.cell(row=row + sector.index(c), column=month_names.index(m) + 4).fill = color_cell
            ws.cell(row=row + sector.index(c), column=month_names.index(m) + 4).font = text_color
            ws.cell(row=row + sector.index(c), column=month_names.index(m) + 4).value = "{:.2f}".format(
                round(decimal.Decimal(sum_total_year), 2))
            total_total += sum_total_year
        color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        text_color = Font(name='Calibri', size=12, bold=True)
        border_cell = Border(left=Side(border_style="thin"),
                             right=Side(border_style="thin"),
                             top=Side(border_style="thin"),
                             bottom=Side(border_style="thin"))
        align_cell = Alignment(horizontal="center")
        ws.cell(row=row + sector.index(c) + 1, column=2).alignment = align_cell
        ws.cell(row=row + sector.index(c) + 1, column=2).border = border_cell
        ws.cell(row=row + sector.index(c) + 1, column=2).fill = color_cell
        ws.cell(row=row + sector.index(c) + 1, column=2).font = text_color
        ws.cell(row=row + sector.index(c) + 1, column=2).value = "TOTAL MES"
        align_cell = Alignment(horizontal="right")
        for t in sum_month:
            ws.cell(row=row + sector.index(c) + 1,
                    column=3 + sum_month.index(t)).alignment = align_cell
            ws.cell(row=row + sector.index(c) + 1,
                    column=3 + sum_month.index(t)).border = border_cell
            ws.cell(row=row + sector.index(c) + 1,
                    column=3 + sum_month.index(t)).fill = color_cell
            ws.cell(row=row + sector.index(c) + 1,
                    column=3 + sum_month.index(t)).font = text_color
            ws.cell(row=row + sector.index(c) + 1, column=3 + sum_month.index(t)).value = "{:.2f}".format(
                round(decimal.Decimal(t), 2))
        ws.cell(row=row + sector.index(c) + 1,
                column=4 + sum_month.index(t)).alignment = align_cell
        ws.cell(row=row + sector.index(c) + 1,
                column=4 + sum_month.index(t)).border = border_cell
        ws.cell(row=row + sector.index(c) + 1,
                column=4 + sum_month.index(t)).fill = color_cell
        ws.cell(row=row + sector.index(c) + 1,
                column=4 + sum_month.index(t)).font = text_color
        ws.cell(row=row + sector.index(c) + 1, column=4 + sum_month.index(t)).value = "{:.2f}".format(
            round(decimal.Decimal(total_total), 2))

        nombre_archivo = "{} - {}.xlsx".format("REPORTE DE COMPRAS", str(year))
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def excel_sales_all_subsidiaries(request, init=None, end=None, pk=0, u=0):
    if init and end:
        start_date = datetime.datetime.strptime(init, "%Y-%m-%d")
        formatted_init = start_date.strftime("%d-%m-%Y")
        end_date = datetime.datetime.strptime(end, "%Y-%m-%d")
        formatted_end = end_date.strftime("%d-%m-%Y")
        unit = bool(u)
        subsidiary = int(pk)
        subsidiary_store_set = SubsidiaryStore.objects.filter(category='V')
        orders = Order.objects.filter(subsidiary_store__in=subsidiary_store_set)
        if start_date == end_date:
            if int(subsidiary) > 0:
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary))
                orders = orders.filter(create_at__date=start_date, type__in=['V', 'R'], subsidiary=subsidiary_obj)
            else:
                orders = orders.filter(create_at__date=start_date, type__in=['V', 'R'])
        else:
            if int(subsidiary) > 0:
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary))
                orders = orders.filter(create_at__date__range=[start_date, end_date], type__in=['V', 'R'],
                                       subsidiary=subsidiary_obj)
            else:
                orders = orders.filter(create_at__date__range=[start_date, end_date], type__in=['V', 'R'])

        wb = Workbook()
        bandera = True
        cont = 1
        if bandera:
            ws = wb.active
            ws.title = str("REPORTE DE VENTAS")  # hoja
            bandera = False
        else:
            ws = wb.create_sheet(str("REPORTE DE VENTAS") + str(cont))
        # Crear el título en la hoja

        if unit:
            if int(subsidiary) > 0:
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary))
                ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
                ws['B2'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
                ws['B2'].font = Font(name='Calibri', size=16, bold=True, color='00FFFFFF')
                ws['B2'] = str("REPORTE DE VENTAS POR UNIDADES DE LA SEDE: ") + str(
                    subsidiary_obj.name) + " DEL " + str(
                    formatted_init) + " AL " + str(formatted_end)
            else:
                ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
                ws['B2'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
                ws['B2'].font = Font(name='Calibri', size=16, bold=True, color='00FFFFFF')
                ws['B2'] = str("REPORTE DE VENTAS POR UNIDADES DEL ") + str(formatted_init) + " AL " + str(
                    formatted_end)

            # Cambiar caracteristicas de las celdas
            ws.merge_cells('B2:K2')

            ws.row_dimensions[2].height = 35

            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20

            # Crear la cabecera
            ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['B4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['B4'] = 'SEDE'

            ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['C4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['C4'] = 'TIPO'

            ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['D4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['D4'] = 'CLIENTE'

            ws['E4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['E4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['E4'] = 'USUARIO'

            ws['F4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['F4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['F4'] = 'TOTAL'

            ws['G4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['G4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['G4'] = 'FECHA'

            ws['H4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['H4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['H4'] = 'BALON DE 10 KG'

            ws['I4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['I4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['I4'] = 'BALON DE 5 KG'

            ws['J4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['J4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['J4'] = 'BALON DE 45 KG'

            ws['K4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['K4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['K4'] = 'BALON DE 15 KG'
            ws.row_dimensions[4].height = 35
            detail = excel_order_by_units(orders)
            row = 5
            for p in detail[0]:
                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=12, bold=False)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="left", vertical="center")
                ws.cell(row=row, column=2).alignment = align_cell
                ws.cell(row=row, column=2).border = border_cell
                ws.cell(row=row, column=2).fill = color_cell
                ws.cell(row=row, column=2).font = text_color
                ws.cell(row=row, column=2).value = p['subsidiary']

                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=12, bold=False)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row, column=3).alignment = align_cell
                ws.cell(row=row, column=3).border = border_cell
                ws.cell(row=row, column=3).fill = color_cell
                ws.cell(row=row, column=3).font = text_color
                ws.cell(row=row, column=3).value = p['type']

                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=12, bold=False)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="left", vertical="center")
                ws.cell(row=row, column=4).alignment = align_cell
                ws.cell(row=row, column=4).border = border_cell
                ws.cell(row=row, column=4).fill = color_cell
                ws.cell(row=row, column=4).font = text_color
                ws.cell(row=row, column=4).value = p['client']
                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=12, bold=False)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="center", vertical="center")

                ws.cell(row=row, column=5).alignment = align_cell
                ws.cell(row=row, column=5).border = border_cell
                ws.cell(row=row, column=5).fill = color_cell
                ws.cell(row=row, column=5).font = text_color
                ws.cell(row=row, column=5).value = p['user']

                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=12, bold=False)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=6).alignment = align_cell
                ws.cell(row=row, column=6).border = border_cell
                ws.cell(row=row, column=6).fill = color_cell
                ws.cell(row=row, column=6).font = text_color
                ws.cell(row=row, column=6).value = "{:.2f}".format(round(p['total'], 2))

                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=12, bold=False)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row, column=7).alignment = align_cell
                ws.cell(row=row, column=7).border = border_cell
                ws.cell(row=row, column=7).fill = color_cell
                ws.cell(row=row, column=7).font = text_color
                ws.cell(row=row, column=7).value = p['create_at'].strftime("%d-%m-%Y")
                column = 0
                for d in p['product_dict']:
                    if d['pk'] == 1:
                        color_cell = PatternFill(start_color='0C66A7', end_color='0C66A7', fill_type="solid")
                    elif d['pk'] == 2:
                        color_cell = PatternFill(start_color='04762D', end_color='04762D', fill_type="solid")
                    elif d['pk'] == 3:
                        color_cell = PatternFill(start_color='9D170A', end_color='9D170A', fill_type="solid")

                    text_color = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                    border_cell = Border(left=Side(border_style="thin"),
                                         right=Side(border_style="thin"),
                                         top=Side(border_style="thin"),
                                         bottom=Side(border_style="thin"))
                    align_cell = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=row, column=8 + column).alignment = align_cell
                    ws.cell(row=row, column=8 + column).border = border_cell
                    ws.cell(row=row, column=8 + column).fill = color_cell
                    ws.cell(row=row, column=8 + column).font = text_color
                    ws.cell(row=row, column=8 + column).value = d['sum']
                    column += 1
                row += 1
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            text_color = Font(name='Calibri', size=12, bold=True)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=5).alignment = align_cell
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = "SUMA TOTAL"

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = detail[1]

            ws.cell(row=row, column=7).alignment = align_cell
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = "SUMA BALONES"

            ws.cell(row=row, column=8).alignment = align_cell
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = str(int(detail[2])) + "[10 KG]"

            ws.cell(row=row, column=9).alignment = align_cell
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = str(int(detail[3])) + "[5 KG]"

            ws.cell(row=row, column=10).alignment = align_cell
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = str(int(detail[4])) + "[45 KG]"

            ws.cell(row=row, column=11).alignment = align_cell
            ws.cell(row=row, column=11).border = border_cell
            ws.cell(row=row, column=11).fill = color_cell
            ws.cell(row=row, column=11).font = text_color
            ws.cell(row=row, column=11).value = str(int(detail[5])) + "[15 KG]"

        else:
            ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B2'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['B2'].font = Font(name='Calibri', size=16, bold=True, color='00FFFFFF')
            ws['B2'] = str("REPORTE DE VENTAS DE ") + str(init) + " AL " + str(end)

            # Cambiar caracteristicas de las celdas
            ws.merge_cells('B2:M2')

            ws.row_dimensions[2].height = 35

            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 25
            ws.column_dimensions['K'].width = 25
            ws.column_dimensions['L'].width = 25
            ws.column_dimensions['M'].width = 25

            # Crear la cabecera
            ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['B4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['B4'] = 'SEDE'

            ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['C4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['C4'] = 'TIPO'

            ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['D4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['D4'] = 'CLIENTE'

            ws['E4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['E4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['E4'] = 'USUARIO'

            ws['F4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['F4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['F4'] = 'TOTAL'

            ws['G4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G4'].fill = PatternFill(start_color='0380D2', end_color='0380D2', fill_type="solid")
            ws['G4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['G4'] = 'FECHA'
            ws['H4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H4'].fill = PatternFill(start_color='028108', end_color='028108', fill_type="solid")
            ws['H4'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['H4'] = 'DETALLES'
            ws.merge_cells('H4:M4')
            ws.row_dimensions[4].height = 30

            ws['B5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B5'].fill = PatternFill(start_color='AD4104', end_color='AD4104', fill_type="solid")
            ws['B5'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['B5'] = 'DATOS GENERALES'
            ws.merge_cells('B5:G5')

            ws['H5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H5'].fill = PatternFill(start_color='5FAE62', end_color='5FAE62', fill_type="solid")
            ws['H5'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['H5'] = 'ID'

            ws['I5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I5'].fill = PatternFill(start_color='4799E7', end_color='4799E7', fill_type="solid")
            ws['I5'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['I5'] = 'PRODUCTO'

            ws['J5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J5'].fill = PatternFill(start_color='A4444C', end_color='A4444C', fill_type="solid")
            ws['J5'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['J5'] = 'UNIDAD'

            ws['K5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K5'].fill = PatternFill(start_color='2F78C0', end_color='2F78C0', fill_type="solid")
            ws['K5'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['K5'] = 'CANTIDAD'

            ws['L5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L5'].fill = PatternFill(start_color='05A04B', end_color='05A04B', fill_type="solid")
            ws['L5'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['L5'] = 'PRECIO'

            ws['M5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['M5'].fill = PatternFill(start_color='0767A9', end_color='0767A9', fill_type="solid")
            ws['M5'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['M5'] = 'SUBTOTAL'
            counter_for = 0
            row = 6
            total = decimal.Decimal(0.00)
            if orders:
                for o in orders:
                    count = o.orderdetail_set.count()

                    color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                    text_color = Font(name='Calibri', size=12, bold=False)
                    border_cell = Border(left=Side(border_style="thin"),
                                         right=Side(border_style="thin"),
                                         top=Side(border_style="thin"),
                                         bottom=Side(border_style="thin"))
                    align_cell = Alignment(horizontal="left", vertical="center")
                    ws.cell(row=row, column=2).alignment = align_cell
                    ws.cell(row=row, column=2).border = border_cell
                    ws.cell(row=row, column=2).fill = color_cell
                    ws.cell(row=row, column=2).font = text_color
                    ws.cell(row=row, column=2).value = o.subsidiary_store.subsidiary.name
                    ws.merge_cells('B{}:B{}'.format(row, row + count - 1))

                    ws.cell(row=row, column=3).alignment = align_cell
                    ws.cell(row=row, column=3).border = border_cell
                    ws.cell(row=row, column=3).fill = color_cell
                    ws.cell(row=row, column=3).font = text_color
                    ws.cell(row=row, column=3).value = o.get_type_display()
                    ws.merge_cells('C{}:C{}'.format(row, row + count - 1))

                    ws.cell(row=row, column=4).alignment = align_cell
                    ws.cell(row=row, column=4).border = border_cell
                    ws.cell(row=row, column=4).fill = color_cell
                    ws.cell(row=row, column=4).font = text_color
                    ws.cell(row=row, column=4).value = o.client.names
                    ws.merge_cells('D{}:D{}'.format(row, row + count - 1))

                    ws.cell(row=row, column=5).alignment = align_cell
                    ws.cell(row=row, column=5).border = border_cell
                    ws.cell(row=row, column=5).fill = color_cell
                    ws.cell(row=row, column=5).font = text_color
                    ws.cell(row=row, column=5).value = o.user.worker_set.last().employee.names
                    ws.merge_cells('E{}:E{}'.format(row, row + count - 1))

                    color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                    text_color = Font(name='Calibri', size=12, bold=True)
                    border_cell = Border(left=Side(border_style="thin"),
                                         right=Side(border_style="thin"),
                                         top=Side(border_style="thin"),
                                         bottom=Side(border_style="thin"))
                    align_cell = Alignment(horizontal="right", vertical="center")

                    ws.cell(row=row, column=6).alignment = align_cell
                    ws.cell(row=row, column=6).border = border_cell
                    ws.cell(row=row, column=6).fill = color_cell
                    ws.cell(row=row, column=6).font = text_color
                    ws.cell(row=row, column=6).value = "{:.2f}".format(round(o.total, 2))
                    ws.merge_cells('F{}:F{}'.format(row, row + count - 1))

                    total += decimal.Decimal(o.total)

                    color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                    text_color = Font(name='Calibri', size=12, bold=True)
                    border_cell = Border(left=Side(border_style="thin"),
                                         right=Side(border_style="thin"),
                                         top=Side(border_style="thin"),
                                         bottom=Side(border_style="thin"))
                    align_cell = Alignment(horizontal="center", vertical="center")

                    ws.cell(row=row, column=7).alignment = align_cell
                    ws.cell(row=row, column=7).border = border_cell
                    ws.cell(row=row, column=7).fill = color_cell
                    ws.cell(row=row, column=7).font = text_color
                    ws.cell(row=row, column=7).value = o.create_at.strftime("%d-%m-%Y")
                    ws.merge_cells('G{}:G{}'.format(row, row + count - 1))
                    detail = o.orderdetail_set.all()
                    if detail is not None:
                        for d in detail:
                            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                            text_color = Font(name='Calibri', size=12, bold=True)
                            border_cell = Border(left=Side(border_style="thin"),
                                                 right=Side(border_style="thin"),
                                                 top=Side(border_style="thin"),
                                                 bottom=Side(border_style="thin"))
                            align_cell = Alignment(horizontal="center", vertical="center")
                            ws.cell(row=row, column=8).alignment = align_cell
                            ws.cell(row=row, column=8).border = border_cell
                            ws.cell(row=row, column=8).fill = color_cell
                            ws.cell(row=row, column=8).font = text_color
                            ws.cell(row=row, column=8).value = d.id
                            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                            text_color = Font(name='Calibri', size=12, bold=True)
                            border_cell = Border(left=Side(border_style="thin"),
                                                 right=Side(border_style="thin"),
                                                 top=Side(border_style="thin"),
                                                 bottom=Side(border_style="thin"))
                            align_cell = Alignment(horizontal="left", vertical="center")
                            ws.cell(row=row, column=9).alignment = align_cell
                            ws.cell(row=row, column=9).border = border_cell
                            ws.cell(row=row, column=9).fill = color_cell
                            ws.cell(row=row, column=9).font = text_color
                            ws.cell(row=row, column=9).value = d.product.name

                            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                            text_color = Font(name='Calibri', size=12, bold=True)
                            border_cell = Border(left=Side(border_style="thin"),
                                                 right=Side(border_style="thin"),
                                                 top=Side(border_style="thin"),
                                                 bottom=Side(border_style="thin"))
                            align_cell = Alignment(horizontal="center", vertical="center")

                            ws.cell(row=row, column=10).alignment = align_cell
                            ws.cell(row=row, column=10).border = border_cell
                            ws.cell(row=row, column=10).fill = color_cell
                            ws.cell(row=row, column=10).font = text_color
                            ws.cell(row=row, column=10).value = d.unit.name

                            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                            text_color = Font(name='Calibri', size=12, bold=True)
                            border_cell = Border(left=Side(border_style="thin"),
                                                 right=Side(border_style="thin"),
                                                 top=Side(border_style="thin"),
                                                 bottom=Side(border_style="thin"))
                            align_cell = Alignment(horizontal="right", vertical="center")

                            ws.cell(row=row, column=11).alignment = align_cell
                            ws.cell(row=row, column=11).border = border_cell
                            ws.cell(row=row, column=11).fill = color_cell
                            ws.cell(row=row, column=11).font = text_color
                            ws.cell(row=row, column=11).value = "{:.2f}".format(
                                round(decimal.Decimal(d.quantity_sold), 2))

                            ws.cell(row=row, column=12).alignment = align_cell
                            ws.cell(row=row, column=12).border = border_cell
                            ws.cell(row=row, column=12).fill = color_cell
                            ws.cell(row=row, column=12).font = text_color
                            ws.cell(row=row, column=12).value = "{:.2f}".format(round(decimal.Decimal(d.price_unit), 2))

                            ws.cell(row=row, column=13).alignment = align_cell
                            ws.cell(row=row, column=13).border = border_cell
                            ws.cell(row=row, column=13).fill = color_cell
                            ws.cell(row=row, column=13).font = text_color
                            ws.cell(row=row, column=13).value = "{:.2f}".format(round(decimal.Decimal(d.multiply()), 2))
                            row = row + 1
                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=12, bold=True)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=4).alignment = align_cell
                ws.cell(row=row, column=4).border = border_cell
                ws.cell(row=row, column=4).fill = color_cell
                ws.cell(row=row, column=4).font = text_color
                ws.cell(row=row, column=4).value = "SUMA TOTAL"

                ws.cell(row=row, column=5).alignment = align_cell
                ws.cell(row=row, column=5).border = border_cell
                ws.cell(row=row, column=5).fill = color_cell
                ws.cell(row=row, column=5).font = text_color
                ws.cell(row=row, column=5).value = "{:.2f}".format(round(total, 2))

        nombre_archivo = "{} DEL {} AL {}.xlsx".format("REPORTE DE VENTAS", str(init), str(end))
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def remove_timezone(dt):
    # Eliminar la información de zona horaria si existe
    if dt is not None and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt


class FilterProduct(TemplateView):

    def get(self, request, *args, **kwargs):
        query = Product.objects.filter(stock__lte=F('minimum'))
        detail = request.GET
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = 'STOCK DE PRODUCTOS' + str(cont)  # hoja
            bandera = False
        else:
            ws = wb.create_sheet('STOCK DE PRODUCTOS' + str(cont))
        # Crear el título en la hoja
        my_date = datetime.datetime.now()
        date_now = my_date.strftime("%d-%m-%Y")
        hour_now = my_date.strftime("%H:%M:%S")
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = 'PRODUCTOS CON STOCK AL MINIMO ' + ' FECHA: ' + str(date_now) + ' HORA: ' + str(hour_now)

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:N1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 9
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 25

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'CODIGO'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'NOMBRE PRODUCTO'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'DESCRIPCION PRODUCTO'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'MARCA'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'FAMILIA'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'TIPO'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'ANCHO'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'LARGO'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'ALTO'

        ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K3'] = 'ALMACEN'

        ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L3'] = 'STOCK'

        ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M3'] = 'MINIMO'

        ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['N3'] = 'PRODUCTO RELACIONADO'

        for p in query.order_by('id'):
            # Pintamos los datos en el reporte
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            stock_cell = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")
            text_color = Font(name='Calibri', size=9)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="center")
            ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).border = border_cell
            ws.cell(row=row, column=2).fill = color_cell
            ws.cell(row=row, column=2).font = text_color
            ws.cell(row=row, column=2).value = p.code

            ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=3).border = border_cell
            ws.cell(row=row, column=3).fill = color_cell
            ws.cell(row=row, column=3).font = text_color
            ws.cell(row=row, column=3).value = p.name

            ws.cell(row=row, column=4).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=4).border = border_cell
            ws.cell(row=row, column=4).fill = color_cell
            ws.cell(row=row, column=4).font = text_color
            ws.cell(row=row, column=4).value = p.description

            ws.cell(row=row, column=5).alignment = align_cell
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = p.brand.name

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = p.family.name

            ws.cell(row=row, column=7).alignment = align_cell
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = p.get_type_display()

            ws.cell(row=row, column=8).alignment = align_cell
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = p.width

            ws.cell(row=row, column=9).alignment = align_cell
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = p.length

            ws.cell(row=row, column=10).alignment = align_cell
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = p.height

            ws.cell(row=row, column=11).alignment = align_cell
            ws.cell(row=row, column=11).border = border_cell
            ws.cell(row=row, column=11).fill = color_cell
            ws.cell(row=row, column=11).font = text_color
            ws.cell(row=row, column=11).value = p.store

            ws.cell(row=row, column=12).alignment = align_cell
            ws.cell(row=row, column=12).border = border_cell
            ws.cell(row=row, column=12).fill = stock_cell
            ws.cell(row=row, column=12).font = text_color
            ws.cell(row=row, column=12).value = p.stock

            ws.cell(row=row, column=13).alignment = align_cell
            ws.cell(row=row, column=13).border = border_cell
            ws.cell(row=row, column=13).fill = color_cell
            ws.cell(row=row, column=13).font = text_color
            ws.cell(row=row, column=13).value = p.minimum

            ws.cell(row=row, column=14).alignment = align_cell
            ws.cell(row=row, column=14).border = border_cell
            ws.cell(row=row, column=14).fill = color_cell
            ws.cell(row=row, column=14).font = text_color
            ws.cell(row=row, column=14).value = p.relation

            cont += 1
            row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "MinStock.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


# class ReportProduct(TemplateView):
#     def get(self, request, *args, **kwargs):
#         # campo = int(request.GET.get('campo'))
#         query = Product.objects.filter(id__lte=10)
#         wb = Workbook()
#         bandera = True
#         cont = 1
#         controlador = 4
#         for q in query:
#             if bandera:
#                 ws = wb.active
#                 ws.title = 'Hoja' + str(cont)
#                 bandera = False
#             else:
#                 ws = wb.create_sheet('Hoja' + str(cont))
#             # Crear el título en la hoja
#             ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
#             ws['B1'].font = Font(name='Calibri', size=12, bold=True)
#             ws['B1'] = 'REPORTE PERSONALIZADO EN EXCEL CON DJANGO'
#
#             # Cambiar caracteristicas de las celdas
#             ws.merge_cells('B1:E1')
#
#             ws.row_dimensions[1].height = 25
#
#             ws.column_dimensions['B'].width = 20
#             ws.column_dimensions['C'].width = 20
#             ws.column_dimensions['D'].width = 20
#             ws.column_dimensions['E'].width = 20
#
#             # Crear la cabecera
#             ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['B3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
#             ws['B3'].font = Font(name='Calibro', size=10, bold=True)
#             ws['B3'] = 'Nombres'
#
#             ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['C3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
#             ws['C3'].font = Font(name='Calibro', size=10, bold=True)
#             ws['C3'] = 'Apellidos'
#
#             ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['D3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
#             ws['D3'].font = Font(name='Calibro', size=10, bold=True)
#             ws['D3'] = 'Dirección'
#
#             ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['E3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
#             ws['E3'].font = Font(name='Calibro', size=10, bold=True)
#             ws['E3'] = 'Edad'
#
#             # Pintamos los datos en el reporte
#             ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
#             ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
#                                                                right=Side(border_style="thin"),
#                                                                top=Side(border_style="thin"),
#                                                                bottom=Side(border_style="thin"))
#             ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
#             ws.cell(row=controlador, column=2).value = q.name
#
#             ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
#             ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
#                                                                right=Side(border_style="thin"),
#                                                                top=Side(border_style="thin"),
#                                                                bottom=Side(border_style="thin"))
#             ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
#             ws.cell(row=controlador, column=3).value = q.code
#
#             ws.cell(row=controlador, column=4).alignment = Alignment(horizontal="center")
#             ws.cell(row=controlador, column=4).border = Border(left=Side(border_style="thin"),
#                                                                right=Side(border_style="thin"),
#                                                                top=Side(border_style="thin"),
#                                                                bottom=Side(border_style="thin"))
#             ws.cell(row=controlador, column=4).font = Font(name='Calibri', size=8)
#             ws.cell(row=controlador, column=4).value = q.brand.name
#
#             ws.cell(row=controlador, column=5).alignment = Alignment(horizontal="center")
#             ws.cell(row=controlador, column=5).border = Border(left=Side(border_style="thin"),
#                                                                right=Side(border_style="thin"),
#                                                                top=Side(border_style="thin"),
#                                                                bottom=Side(border_style="thin"))
#             ws.cell(row=controlador, column=5).font = Font(name='Calibri', size=8)
#             ws.cell(row=controlador, column=5).value = q.family.name
#
#             cont += 1
#
#         # Establecer el nombre de mi archivo
#         nombre_archivo = "Productos.xlsx"
#         # Definir el tipo de respuesta que se va a dar
#         response = HttpResponse(content_type="application/ms-excel")
#         contenido = "attachment; filename = {0}".format(nombre_archivo)
#         response["Content-Disposition"] = contenido
#         wb.save(response)
#         return response
def purchase_excel(request, init=None, end=None):
    if init and end:
        order_set = Order.objects.filter(create_at__range=(init, end), type='C').order_by('id')
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = str("Reporte-compras")  # hoja
            bandera = False
        else:
            ws = wb.create_sheet(str("Reporte-compras") + str(cont))
        # Crear el título en la hoja
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = str("REPORTE DE COMPRAS")
        # " " + str(datetime.date(init).strptime("%d-%m-%Y")) + "-" + str(datetime.date(end).strftime("%d-%m-%Y"))
        # str(end)

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:W1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 35
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 15

        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 40
        ws.column_dimensions['S'].width = 12
        ws.column_dimensions['T'].width = 12
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'Nº'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'CODIGO'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'COMPROBANTE'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'TIPO'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'ESTADO'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'FECHA'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'INGRESO'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'DOCUMENTO'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'PROVEEDOR'

        ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K3'] = 'TOTAL'

        ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L3'] = 'MONEDA'

        ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M3'] = 'CAMBIO'

        ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['N3'] = 'SUBTOTAL(S/.)'

        ws['O3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['O3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['O3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['O3'] = 'IGV(S/.)'

        ws['P3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['P3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['P3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['P3'] = 'TOTAL(S/.)'

        ws['Q3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['Q3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['Q3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['Q3'] = 'CODIGO'

        ws['R3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['R3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['R3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['R3'] = 'PRODUCTO'

        ws['S3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['S3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['S3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['S3'] = 'MEDIDA'

        ws['T3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['T3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['T3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['T3'] = 'CANTIDAD'

        ws['U3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['U3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['U3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['U3'] = 'UNIDAD'

        ws['V3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['V3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['V3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['V3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['V3'] = 'PRECIO'

        ws['W3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['W3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['W3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['W3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['W3'] = 'TOTAL'
        total = decimal.Decimal(0.000000)
        subtotal = decimal.Decimal(0.000000)
        total_igv = decimal.Decimal(0.000000)
        for o in order_set:
            # Pintamos los datos en el reporteç
            date_one = "-"
            if o.create_at:
                date_one = o.create_at.strftime("%d-%m-%Y")
            date_two = "-"
            if o.invoice_date:
                date_two = o.invoice_date.strftime("%d-%m-%Y")
            date_three = "-"
            if o.date_document:
                date_three = o.date_document.strftime("%d-%m-%Y")
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            text_color = Font(name='Calibri', size=9)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(vertical="center", horizontal="center")

            # ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).border = border_cell
            ws.cell(row=row, column=2).fill = color_cell
            ws.cell(row=row, column=2).font = text_color
            ws.cell(row=row, column=2).value = o.number

            ws.cell(row=row, column=3).alignment = align_cell
            ws.cell(row=row, column=3).border = border_cell
            ws.cell(row=row, column=3).fill = color_cell
            ws.cell(row=row, column=3).font = text_color
            ws.cell(row=row, column=3).value = o.get_code()

            ws.cell(row=row, column=4).alignment = align_cell
            ws.cell(row=row, column=4).border = border_cell
            ws.cell(row=row, column=4).fill = color_cell
            ws.cell(row=row, column=4).font = text_color
            ws.cell(row=row, column=4).value = o.invoice_number

            ws.cell(row=row, column=5).alignment = align_cell
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = o.get_doc_display()

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = o.get_status_display()

            ws.cell(row=row, column=7).alignment = align_cell
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = date_one

            ws.cell(row=row, column=8).alignment = align_cell
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = date_two

            ws.cell(row=row, column=9).alignment = align_cell
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = date_three

            ws.cell(row=row, column=10).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = o.person.names

            ws.cell(row=row, column=11).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=11).border = border_cell
            ws.cell(row=row, column=11).fill = color_cell
            ws.cell(row=row, column=11).font = text_color
            ws.cell(row=row, column=11).value = o.total

            ws.cell(row=row, column=12).alignment = align_cell
            ws.cell(row=row, column=12).border = border_cell
            ws.cell(row=row, column=12).fill = color_cell
            ws.cell(row=row, column=12).font = text_color
            ws.cell(row=row, column=12).value = o.get_coin_display()

            ws.cell(row=row, column=13).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=13).border = border_cell
            ws.cell(row=row, column=13).fill = color_cell
            ws.cell(row=row, column=13).font = text_color
            ws.cell(row=row, column=13).value = o.change

            amount = o.change * o.total
            amount_sin_igv = amount / decimal.Decimal(1.1800)
            igv = amount - (amount / decimal.Decimal(1.1800))

            ws.cell(row=row, column=14).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=14).border = border_cell
            ws.cell(row=row, column=14).fill = color_cell
            ws.cell(row=row, column=14).font = text_color
            ws.cell(row=row, column=14).value = round(amount_sin_igv, 2)

            ws.cell(row=row, column=15).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=15).border = border_cell
            ws.cell(row=row, column=15).fill = color_cell
            ws.cell(row=row, column=15).font = text_color
            ws.cell(row=row, column=15).value = round(igv, 2)

            ws.cell(row=row, column=16).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=16).border = border_cell
            ws.cell(row=row, column=16).fill = color_cell
            ws.cell(row=row, column=16).font = text_color
            ws.cell(row=row, column=16).value = round(amount, 2)
            subtotal = subtotal + amount_sin_igv
            total_igv = total_igv + igv
            total = total + amount
            di = 0
            for d in o.orderdetail_set.filter(is_state=True).order_by('id'):
                di += 1
                ws.cell(row=row, column=17).alignment = align_cell
                ws.cell(row=row, column=17).border = border_cell
                ws.cell(row=row, column=17).fill = color_cell
                ws.cell(row=row, column=17).font = text_color
                ws.cell(row=row, column=17).value = d.product.code

                ws.cell(row=row, column=18).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=row, column=18).border = border_cell
                ws.cell(row=row, column=18).fill = color_cell
                ws.cell(row=row, column=18).font = text_color
                ws.cell(row=row, column=18).value = d.product.name

                ws.cell(row=row, column=19).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=19).border = border_cell
                ws.cell(row=row, column=19).fill = color_cell
                ws.cell(row=row, column=19).font = text_color
                ws.cell(row=row, column=19).value = d.product.measure()

                ws.cell(row=row, column=20).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=20).border = border_cell
                ws.cell(row=row, column=20).fill = color_cell
                ws.cell(row=row, column=20).font = text_color
                ws.cell(row=row, column=20).value = d.quantity

                ws.cell(row=row, column=21).alignment = align_cell
                ws.cell(row=row, column=21).border = border_cell
                ws.cell(row=row, column=21).fill = color_cell
                ws.cell(row=row, column=21).font = text_color
                ws.cell(row=row, column=21).value = d.get_unit_display()

                ws.cell(row=row, column=22).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=22).border = border_cell
                ws.cell(row=row, column=22).fill = color_cell
                ws.cell(row=row, column=22).font = text_color
                ws.cell(row=row, column=22).value = d.price

                ws.cell(row=row, column=23).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=23).border = border_cell
                ws.cell(row=row, column=23).fill = color_cell
                ws.cell(row=row, column=23).font = text_color
                ws.cell(row=row, column=23).value = d.amount()
                row += 1
            # ws.merge_cells('B1:S1')
            cont += 1
            ws.merge_cells('B' + str(row - di) + ':B' + str(row - 1))
            ws.merge_cells('C' + str(row - di) + ':C' + str(row - 1))
            ws.merge_cells('D' + str(row - di) + ':D' + str(row - 1))
            ws.merge_cells('E' + str(row - di) + ':E' + str(row - 1))
            ws.merge_cells('F' + str(row - di) + ':F' + str(row - 1))
            ws.merge_cells('G' + str(row - di) + ':G' + str(row - 1))
            ws.merge_cells('H' + str(row - di) + ':H' + str(row - 1))
            ws.merge_cells('I' + str(row - di) + ':I' + str(row - 1))
            ws.merge_cells('J' + str(row - di) + ':J' + str(row - 1))
            ws.merge_cells('K' + str(row - di) + ':K' + str(row - 1))
            ws.merge_cells('L' + str(row - di) + ':L' + str(row - 1))
            ws.merge_cells('M' + str(row - di) + ':M' + str(row - 1))
            ws.merge_cells('N' + str(row - di) + ':N' + str(row - 1))
            ws.merge_cells('O' + str(row - di) + ':O' + str(row - 1))
            ws.merge_cells('P' + str(row - di) + ':P' + str(row - 1))

        ws.cell(row=row, column=14).alignment = Alignment(vertical="center", horizontal="right")
        ws.cell(row=row, column=14).border = Border(left=Side(border_style="thin"),
                                                    right=Side(border_style="thin"),
                                                    top=Side(border_style="thin"),
                                                    bottom=Side(border_style="thin"))
        ws.cell(row=row, column=14).fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        ws.cell(row=row, column=14).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=14).value = round(subtotal, 2)

        ws.cell(row=row, column=15).alignment = Alignment(vertical="center", horizontal="right")
        ws.cell(row=row, column=15).border = Border(left=Side(border_style="thin"),
                                                    right=Side(border_style="thin"),
                                                    top=Side(border_style="thin"),
                                                    bottom=Side(border_style="thin"))
        ws.cell(row=row, column=15).fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        ws.cell(row=row, column=15).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=15).value = round(total_igv, 2)

        ws.cell(row=row, column=16).alignment = Alignment(vertical="center", horizontal="right")
        ws.cell(row=row, column=16).border = Border(left=Side(border_style="thin"),
                                                    right=Side(border_style="thin"),
                                                    top=Side(border_style="thin"),
                                                    bottom=Side(border_style="thin"))
        ws.cell(row=row, column=16).fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        ws.cell(row=row, column=16).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=16).value = round(total, 2)
        # row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "Compras-jhunior.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


# def export_product_filter(request, c1=None, c2=None, c3=None, c4=None, array_id=None):
#     product_set = []
#     _arr = []
#     if array_id is not None and array_id != " ":
#         str1 = array_id.replace(']', '').replace('[', '')
#         _arr = str1.replace('"', '').replace("'", '').split(",")
#     if c1 != "0" and len(c1) > 3:
#         product_set = Product.objects.filter(name__contains=c1.upper()).order_by('id')
#     elif c4 != "0":
#         product_set = Product.objects.filter(code=int(c4))
#     elif c2 != "0":
#         if c3 == '1':
#             product_set = Product.objects.filter(
#                 brand__name__contains=c2.upper()).select_related('brand', 'family').prefetch_related(
#                 Prefetch(
#                     'presentation_set', queryset=Presentation.objects.select_related('product')
#                 )
#             ).order_by('id')
#         elif c3 == '2':
#             product_set = Product.objects.filter(family__name__contains=c2.upper()).order_by('id')
#         else:
#             product_set = []
#     elif c3 == '1':
#         product_set = Product.objects.all().select_related('brand', 'family').prefetch_related(
#             Prefetch(
#                 'presentation_set', queryset=Presentation.objects.select_related('product')
#             )
#         ).order_by('id')
#     else:
#         product_set = Product.objects.filter(id__in=_arr).order_by('id')
#
#     wb = Workbook()
#     bandera = True
#     cont = 1
#     row = 4
#     if bandera:
#         ws = wb.active
#         ws.title = str("Productos Filtrados")  # hoja
#         bandera = False
#     else:
#         ws = wb.create_sheet(str("Productos Filtrados") + str(cont))
#     # Crear el título en la hoja
#     ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['B1'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
#     ws['B1'] = 'LISTA DE PRODUCTOS'
#
#     # Cambiar caracteristicas de las celdas
#     ws.merge_cells('B1:R1')
#
#     ws.row_dimensions[1].height = 25
#
#     ws.column_dimensions['B'].width = 9
#     ws.column_dimensions['C'].width = 50
#     ws.column_dimensions['D'].width = 50
#     ws.column_dimensions['E'].width = 9
#     ws.column_dimensions['F'].width = 9
#     ws.column_dimensions['G'].width = 9
#     ws.column_dimensions['H'].width = 9
#     ws.column_dimensions['I'].width = 9
#     ws.column_dimensions['J'].width = 9
#     ws.column_dimensions['K'].width = 20
#     ws.column_dimensions['L'].width = 9
#     ws.column_dimensions['M'].width = 9
#     ws.column_dimensions['N'].width = 25
#     ws.column_dimensions['O'].width = 15
#     ws.column_dimensions['P'].width = 20
#     ws.column_dimensions['Q'].width = 20
#     ws.column_dimensions['R'].width = 15
#
#     # Crear la cabecera
#     ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['B3'] = 'CODIGO'
#
#     ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['C3'] = 'NOMBRE PRODUCTO'
#
#     ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['D3'] = 'DESCRIPCION PRODUCTO'
#
#     ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['E3'] = 'MARCA'
#
#     ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['F3'] = 'FAMILIA'
#
#     ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['G3'] = 'TIPO'
#
#     ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['H3'] = 'ANCHO'
#
#     ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['I3'] = 'LARGO'
#
#     ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['J3'] = 'ALTO'
#
#     ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['K3'] = 'ALMACEN'
#
#     ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['L3'] = 'STOCK'
#
#     ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['M3'] = 'MINIMO'
#
#     ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['N3'] = 'PRODUCTO RELACIONADO'
#
#     ws['O3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['O3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['O3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['O3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['O3'] = 'CANTIDAD'
#
#     ws['P3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['P3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['P3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['P3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['P3'] = 'UNIDAD'
#
#     ws['Q3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['Q3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['Q3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['Q3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['Q3'] = 'CANTIDAD UNITARIA'
#
#     ws['R3'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['R3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#     ws['R3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#     ws['R3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#     ws['R3'] = 'PRICE'
#
#     for p in product_set:
#         # Pintamos los datos en el reporte
#         color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
#         stock_cell = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")
#         text_color = Font(name='Calibri', size=9)
#         border_cell = Border(left=Side(border_style="thin"),
#                              right=Side(border_style="thin"),
#                              top=Side(border_style="thin"),
#                              bottom=Side(border_style="thin"))
#         align_cell = Alignment(horizontal="center")
#         ws.cell(row=row, column=2).alignment = align_cell
#         ws.cell(row=row, column=2).border = border_cell
#         ws.cell(row=row, column=2).fill = color_cell
#         ws.cell(row=row, column=2).font = text_color
#         ws.cell(row=row, column=2).value = p.code
#
#         ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
#         ws.cell(row=row, column=3).border = border_cell
#         ws.cell(row=row, column=3).fill = color_cell
#         ws.cell(row=row, column=3).font = text_color
#         ws.cell(row=row, column=3).value = p.name
#
#         ws.cell(row=row, column=4).alignment = Alignment(horizontal="left")
#         ws.cell(row=row, column=4).border = border_cell
#         ws.cell(row=row, column=4).fill = color_cell
#         ws.cell(row=row, column=4).font = text_color
#         ws.cell(row=row, column=4).value = p.description
#
#         ws.cell(row=row, column=5).alignment = align_cell
#         ws.cell(row=row, column=5).border = border_cell
#         ws.cell(row=row, column=5).fill = color_cell
#         ws.cell(row=row, column=5).font = text_color
#         ws.cell(row=row, column=5).value = p.brand.name
#
#         ws.cell(row=row, column=6).alignment = align_cell
#         ws.cell(row=row, column=6).border = border_cell
#         ws.cell(row=row, column=6).fill = color_cell
#         ws.cell(row=row, column=6).font = text_color
#         ws.cell(row=row, column=6).value = p.family.name
#
#         ws.cell(row=row, column=7).alignment = align_cell
#         ws.cell(row=row, column=7).border = border_cell
#         ws.cell(row=row, column=7).fill = color_cell
#         ws.cell(row=row, column=7).font = text_color
#         ws.cell(row=row, column=7).value = p.get_type_display()
#
#         ws.cell(row=row, column=8).alignment = align_cell
#         ws.cell(row=row, column=8).border = border_cell
#         ws.cell(row=row, column=8).fill = color_cell
#         ws.cell(row=row, column=8).font = text_color
#         ws.cell(row=row, column=8).value = p.width
#
#         ws.cell(row=row, column=9).alignment = align_cell
#         ws.cell(row=row, column=9).border = border_cell
#         ws.cell(row=row, column=9).fill = color_cell
#         ws.cell(row=row, column=9).font = text_color
#         ws.cell(row=row, column=9).value = p.length
#
#         ws.cell(row=row, column=10).alignment = align_cell
#         ws.cell(row=row, column=10).border = border_cell
#         ws.cell(row=row, column=10).fill = color_cell
#         ws.cell(row=row, column=10).font = text_color
#         ws.cell(row=row, column=10).value = p.height
#
#         ws.cell(row=row, column=11).alignment = align_cell
#         ws.cell(row=row, column=11).border = border_cell
#         ws.cell(row=row, column=11).fill = color_cell
#         ws.cell(row=row, column=11).font = text_color
#         ws.cell(row=row, column=11).value = p.store
#
#         ws.cell(row=row, column=12).alignment = align_cell
#         ws.cell(row=row, column=12).border = border_cell
#         ws.cell(row=row, column=12).fill = stock_cell
#         ws.cell(row=row, column=12).font = text_color
#         ws.cell(row=row, column=12).value = p.stock
#
#         ws.cell(row=row, column=13).alignment = align_cell
#         ws.cell(row=row, column=13).border = border_cell
#         ws.cell(row=row, column=13).fill = color_cell
#         ws.cell(row=row, column=13).font = text_color
#         ws.cell(row=row, column=13).value = p.minimum
#
#         ws.cell(row=row, column=14).alignment = align_cell
#         ws.cell(row=row, column=14).border = border_cell
#         ws.cell(row=row, column=14).fill = color_cell
#         ws.cell(row=row, column=14).font = text_color
#         ws.cell(row=row, column=14).value = p.relation
#         dp = p.presentation_set.count()
#         df = 0
#
#         for pt in p.presentation_set.all():
#             df += 1
#             ws.cell(row=row, column=15).alignment = align_cell
#             ws.cell(row=row, column=15).border = border_cell
#             ws.cell(row=row, column=15).fill = color_cell
#             ws.cell(row=row, column=15).font = text_color
#             ws.cell(row=row, column=15).value = pt.quantity
#
#             ws.cell(row=row, column=16).alignment = align_cell
#             ws.cell(row=row, column=16).border = border_cell
#             ws.cell(row=row, column=16).fill = color_cell
#             ws.cell(row=row, column=16).font = text_color
#             ws.cell(row=row, column=16).value = pt.get_unit_display()
#
#             ws.cell(row=row, column=17).alignment = align_cell
#             ws.cell(row=row, column=17).border = border_cell
#             ws.cell(row=row, column=17).fill = color_cell
#             ws.cell(row=row, column=17).font = text_color
#             ws.cell(row=row, column=17).value = pt.quantity_niu
#
#             ws.cell(row=row, column=18).alignment = align_cell
#             ws.cell(row=row, column=18).border = border_cell
#             ws.cell(row=row, column=18).fill = color_cell
#             ws.cell(row=row, column=18).font = text_color
#             ws.cell(row=row, column=18).value = pt.price
#             if df < dp:
#                 row += 1
#         # ws.merge_cells('B' + str(row-df) + ':B' + str(row))
#         cont += 1
#         row += 1
#
#     # Establecer el nombre de mi archivo
#     nombre_archivo = "ProductFilter.xlsx"
#     # Definir el tipo de respuesta que se va a dar
#     response = HttpResponse(content_type="application/ms-excel")
#     contenido = "attachment; filename = {0}".format(nombre_archivo)
#     response["Content-Disposition"] = contenido
#     wb.save(response)
#     return response
#
#
# def report_kardex_cont(request, init=None, end=None, pk=None):
#     if init and end and pk:
#         product_obj = Product.objects.get(id=int(pk))
#         # m = month[5:]
#         detail_set = OrderDetail.objects.filter(product=product_obj, order__type__in=['V', 'C'],
#                                                 order__create_at__range=(init, end)).order_by('id')
#
#         order_detail_obj = OrderDetail.objects.filter(product=product_obj, order__type__in=['V', 'C']).first()
#         wb = Workbook()
#         bandera = True
#         cont = 1
#         row = 4
#         if bandera:
#             ws = wb.active
#             ws.title = str("PRODUCTO")  # hoja
#             bandera = False
#         else:
#             ws = wb.create_sheet(str(product_obj.name) + str(cont))
#         # Crear el título en la hoja
#         my_date = datetime.datetime.now()
#         date_now = my_date.strftime("%d-%m-%Y")
#         hour_now = my_date.strftime("%H:%M:%S")
#         ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['B1'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
#         ws['B1'] = str(product_obj.name.upper())
#
#         # Cambiar caracteristicas de las celdas
#         ws.merge_cells('B1:M1')
#
#         ws.row_dimensions[1].height = 25
#
#         ws.column_dimensions['B'].width = 15
#         ws.column_dimensions['C'].width = 15
#         ws.column_dimensions['D'].width = 15
#         ws.column_dimensions['E'].width = 15
#         ws.column_dimensions['F'].width = 15
#         ws.column_dimensions['G'].width = 15
#         ws.column_dimensions['H'].width = 10
#         ws.column_dimensions['I'].width = 10
#         ws.column_dimensions['J'].width = 15
#         ws.column_dimensions['K'].width = 15
#         ws.column_dimensions['L'].width = 15
#         ws.column_dimensions['M'].width = 15
#
#         # Crear la cabecera
#         ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['B3'] = 'FECHA'
#
#         ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['C3'] = 'ORDEN'
#
#         ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['D3'] = 'OPERACIÓN'
#
#         ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['E3'] = 'CANTIDAD'
#
#         ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['F3'] = 'MEDIDA'
#
#         ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['G3'] = 'ENTRADA'
#
#         ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['H3'] = 'SALIDA'
#
#         ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['I3'] = 'SALDO'
#
#         ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['J3'] = 'INGRESO'
#
#         ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['K3'] = 'EGRESO'
#
#         ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['L3'] = 'SALDO'
#
#         ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
#         ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#         ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
#         ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
#         ws['M3'] = 'PROMEDIO'
#         counter_for = 0
#         for d in detail_set:
#             if counter_for == 0 and order_detail_obj.id == d.id:
#                 color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
#                 text_color = Font(name='Calibri', size=9)
#                 border_cell = Border(left=Side(border_style="thin"),
#                                      right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"),
#                                      bottom=Side(border_style="thin"))
#                 align_cell = Alignment(horizontal="center")
#
#                 ws.cell(row=row, column=2).alignment = align_cell
#                 ws.cell(row=row, column=2).border = border_cell
#                 ws.cell(row=row, column=2).fill = color_cell
#                 ws.cell(row=row, column=2).font = text_color
#                 ws.cell(row=row, column=2).value = "-"
#
#                 ws.cell(row=row, column=3).alignment = align_cell
#                 ws.cell(row=row, column=3).border = border_cell
#                 ws.cell(row=row, column=3).fill = color_cell
#                 ws.cell(row=row, column=3).font = text_color
#                 ws.cell(row=row, column=3).value = "INICIO"
#
#                 ws.cell(row=row, column=4).alignment = align_cell
#                 ws.cell(row=row, column=4).border = border_cell
#                 ws.cell(row=row, column=4).fill = color_cell
#                 ws.cell(row=row, column=4).font = text_color
#                 ws.cell(row=row, column=4).value = "RESTANTE"
#
#                 ws.cell(row=row, column=5).alignment = align_cell
#                 ws.cell(row=row, column=5).border = border_cell
#                 ws.cell(row=row, column=5).fill = color_cell
#                 ws.cell(row=row, column=5).font = text_color
#                 ws.cell(row=row, column=5).value = "-"
#
#                 ws.cell(row=row, column=6).alignment = align_cell
#                 ws.cell(row=row, column=6).border = border_cell
#                 ws.cell(row=row, column=6).fill = color_cell
#                 ws.cell(row=row, column=6).font = text_color
#                 ws.cell(row=row, column=6).value = "UNIDAD"
#
#                 ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
#                 ws.cell(row=row, column=7).border = border_cell
#                 ws.cell(row=row, column=7).fill = color_cell
#                 ws.cell(row=row, column=7).font = text_color
#                 ws.cell(row=row, column=7).value = "-"
#
#                 ws.cell(row=row, column=8).alignment = align_cell
#                 ws.cell(row=row, column=8).border = border_cell
#                 ws.cell(row=row, column=8).fill = color_cell
#                 ws.cell(row=row, column=8).font = text_color
#                 ws.cell(row=row, column=8).value = "-"
#
#                 ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
#                 ws.cell(row=row, column=9).border = border_cell
#                 ws.cell(row=row, column=9).fill = color_cell
#                 ws.cell(row=row, column=9).font = text_color
#                 ws.cell(row=row, column=9).value = round(d.previous(), 2)
#
#                 ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
#                 ws.cell(row=row, column=10).border = border_cell
#                 ws.cell(row=row, column=10).fill = color_cell
#                 ws.cell(row=row, column=10).font = text_color
#                 ws.cell(row=row, column=10).value = "-"
#
#                 ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
#                 ws.cell(row=row, column=11).border = border_cell
#                 ws.cell(row=row, column=11).fill = color_cell
#                 ws.cell(row=row, column=11).font = text_color
#                 ws.cell(row=row, column=11).value = "-"
#
#                 ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
#                 ws.cell(row=row, column=12).border = border_cell
#                 ws.cell(row=row, column=12).fill = color_cell
#                 ws.cell(row=row, column=12).font = text_color
#                 ws.cell(row=row, column=12).value = round(d.total_initial(), 2)
#
#                 ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
#                 ws.cell(row=row, column=13).border = border_cell
#                 ws.cell(row=row, column=13).fill = color_cell
#                 ws.cell(row=row, column=13).font = text_color
#                 ws.cell(row=row, column=13).value = round(d.product.price_unit(), 2)
#                 row += 1
#             # Estado activo
#             if d.is_state:
#                 if d.operation == 'S':
#                     color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
#                     text_color = Font(name='Calibri', size=9)
#                     border_cell = Border(left=Side(border_style="thin"),
#                                          right=Side(border_style="thin"),
#                                          top=Side(border_style="thin"),
#                                          bottom=Side(border_style="thin"))
#                     align_cell = Alignment(horizontal="center")
#
#                     ws.cell(row=row, column=2).alignment = align_cell
#                     ws.cell(row=row, column=2).border = border_cell
#                     ws.cell(row=row, column=2).fill = color_cell
#                     ws.cell(row=row, column=2).font = text_color
#                     ws.cell(row=row, column=2).value = d.order.create_at
#
#                     ws.cell(row=row, column=3).alignment = align_cell
#                     ws.cell(row=row, column=3).border = border_cell
#                     ws.cell(row=row, column=3).fill = color_cell
#                     ws.cell(row=row, column=3).font = text_color
#                     ws.cell(row=row, column=3).value = str(d.order.get_type_display()).upper() + "Nº " + str(
#                         d.order.number)
#
#                     ws.cell(row=row, column=4).alignment = align_cell
#                     ws.cell(row=row, column=4).border = border_cell
#                     ws.cell(row=row, column=4).fill = color_cell
#                     ws.cell(row=row, column=4).font = text_color
#                     ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()
#
#                     ws.cell(row=row, column=5).alignment = align_cell
#                     ws.cell(row=row, column=5).border = border_cell
#                     ws.cell(row=row, column=5).fill = color_cell
#                     ws.cell(row=row, column=5).font = text_color
#                     ws.cell(row=row, column=5).value = str(d.quantity)
#
#                     ws.cell(row=row, column=6).alignment = align_cell
#                     ws.cell(row=row, column=6).border = border_cell
#                     ws.cell(row=row, column=6).fill = color_cell
#                     ws.cell(row=row, column=6).font = text_color
#                     ws.cell(row=row, column=6).value = str(d.get_unit_display())
#
#                     ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=7).border = border_cell
#                     ws.cell(row=row, column=7).fill = color_cell
#                     ws.cell(row=row, column=7).font = text_color
#                     ws.cell(row=row, column=7).value = "-"
#
#                     ws.cell(row=row, column=8).alignment = align_cell
#                     ws.cell(row=row, column=8).border = border_cell
#                     ws.cell(row=row, column=8).fill = color_cell
#                     ws.cell(row=row, column=8).font = text_color
#                     ws.cell(row=row, column=8).value = round(d.quantity_niu, 2)
#
#                     ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=9).border = border_cell
#                     ws.cell(row=row, column=9).fill = color_cell
#                     ws.cell(row=row, column=9).font = text_color
#                     ws.cell(row=row, column=9).value = round(d.quantity_remaining, 2)
#
#                     ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=10).border = border_cell
#                     ws.cell(row=row, column=10).fill = color_cell
#                     ws.cell(row=row, column=10).font = text_color
#                     ws.cell(row=row, column=10).value = "-"
#
#                     ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=11).border = border_cell
#                     ws.cell(row=row, column=11).fill = color_cell
#                     ws.cell(row=row, column=11).font = text_color
#                     ws.cell(row=row, column=11).value = d.amount()
#
#                     ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=12).border = border_cell
#                     ws.cell(row=row, column=12).fill = color_cell
#                     ws.cell(row=row, column=12).font = text_color
#                     ws.cell(row=row, column=12).value = d.balance_remaining()
#
#                     ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=13).border = border_cell
#                     ws.cell(row=row, column=13).fill = color_cell
#                     ws.cell(row=row, column=13).font = text_color
#                     ws.cell(row=row, column=13).value = d.price
#                 elif d.operation == 'E':
#                     color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
#                     text_color = Font(name='Calibri', size=9)
#                     border_cell = Border(left=Side(border_style="thin"),
#                                          right=Side(border_style="thin"),
#                                          top=Side(border_style="thin"),
#                                          bottom=Side(border_style="thin"))
#                     align_cell = Alignment(horizontal="center")
#
#                     ws.cell(row=row, column=2).alignment = align_cell
#                     ws.cell(row=row, column=2).border = border_cell
#                     ws.cell(row=row, column=2).fill = color_cell
#                     ws.cell(row=row, column=2).font = text_color
#                     ws.cell(row=row, column=2).value = d.order.create_at
#
#                     ws.cell(row=row, column=3).alignment = align_cell
#                     ws.cell(row=row, column=3).border = border_cell
#                     ws.cell(row=row, column=3).fill = color_cell
#                     ws.cell(row=row, column=3).font = text_color
#                     ws.cell(row=row, column=3).value = str(d.order.get_type_display()).upper() + "Nº " + str(
#                         d.order.number)
#
#                     ws.cell(row=row, column=4).alignment = align_cell
#                     ws.cell(row=row, column=4).border = border_cell
#                     ws.cell(row=row, column=4).fill = color_cell
#                     ws.cell(row=row, column=4).font = text_color
#                     ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()
#
#                     ws.cell(row=row, column=5).alignment = align_cell
#                     ws.cell(row=row, column=5).border = border_cell
#                     ws.cell(row=row, column=5).fill = color_cell
#                     ws.cell(row=row, column=5).font = text_color
#                     ws.cell(row=row, column=5).value = str(d.quantity)
#
#                     ws.cell(row=row, column=6).alignment = align_cell
#                     ws.cell(row=row, column=6).border = border_cell
#                     ws.cell(row=row, column=6).fill = color_cell
#                     ws.cell(row=row, column=6).font = text_color
#                     ws.cell(row=row, column=6).value = str(d.get_unit_display())
#
#                     ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=7).border = border_cell
#                     ws.cell(row=row, column=7).fill = color_cell
#                     ws.cell(row=row, column=7).font = text_color
#                     ws.cell(row=row, column=7).value = d.quantity_niu
#
#                     ws.cell(row=row, column=8).alignment = align_cell
#                     ws.cell(row=row, column=8).border = border_cell
#                     ws.cell(row=row, column=8).fill = color_cell
#                     ws.cell(row=row, column=8).font = text_color
#                     ws.cell(row=row, column=8).value = "-"
#
#                     ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=9).border = border_cell
#                     ws.cell(row=row, column=9).fill = color_cell
#                     ws.cell(row=row, column=9).font = text_color
#                     ws.cell(row=row, column=9).value = d.quantity_remaining
#
#                     ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=10).border = border_cell
#                     ws.cell(row=row, column=10).fill = color_cell
#                     ws.cell(row=row, column=10).font = text_color
#                     ws.cell(row=row, column=10).value = d.amount()
#
#                     ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=11).border = border_cell
#                     ws.cell(row=row, column=11).fill = color_cell
#                     ws.cell(row=row, column=11).font = text_color
#                     ws.cell(row=row, column=11).value = "-"
#
#                     ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=12).border = border_cell
#                     ws.cell(row=row, column=12).fill = color_cell
#                     ws.cell(row=row, column=12).font = text_color
#                     ws.cell(row=row, column=12).value = d.balance_remaining()
#
#                     ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=13).border = border_cell
#                     ws.cell(row=row, column=13).fill = color_cell
#                     ws.cell(row=row, column=13).font = text_color
#                     ws.cell(row=row, column=13).value = d.price
#             else:
#                 if d.operation == 'S':
#                     color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
#                     text_color = Font(name='Calibri', size=9)
#                     border_cell = Border(left=Side(border_style="thin"),
#                                          right=Side(border_style="thin"),
#                                          top=Side(border_style="thin"),
#                                          bottom=Side(border_style="thin"))
#                     align_cell = Alignment(horizontal="center")
#
#                     ws.cell(row=row, column=2).alignment = align_cell
#                     ws.cell(row=row, column=2).border = border_cell
#                     ws.cell(row=row, column=2).fill = color_cell
#                     ws.cell(row=row, column=2).font = text_color
#                     ws.cell(row=row, column=2).value = d.order.create_at
#
#                     ws.cell(row=row, column=3).alignment = align_cell
#                     ws.cell(row=row, column=3).border = border_cell
#                     ws.cell(row=row, column=3).fill = color_cell
#                     ws.cell(row=row, column=3).font = text_color
#                     ws.cell(row=row, column=3).value = str(d.order.get_type_display()).upper() + "Nº " + str(
#                         d.order.number)
#
#                     ws.cell(row=row, column=4).alignment = align_cell
#                     ws.cell(row=row, column=4).border = border_cell
#                     ws.cell(row=row, column=4).fill = color_cell
#                     ws.cell(row=row, column=4).font = text_color
#                     ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()
#
#                     ws.cell(row=row, column=5).alignment = align_cell
#                     ws.cell(row=row, column=5).border = border_cell
#                     ws.cell(row=row, column=5).fill = color_cell
#                     ws.cell(row=row, column=5).font = text_color
#                     ws.cell(row=row, column=5).value = str(d.quantity)
#
#                     ws.cell(row=row, column=6).alignment = align_cell
#                     ws.cell(row=row, column=6).border = border_cell
#                     ws.cell(row=row, column=6).fill = color_cell
#                     ws.cell(row=row, column=6).font = text_color
#                     ws.cell(row=row, column=6).value = str(d.get_unit_display())
#
#                     ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=7).border = border_cell
#                     ws.cell(row=row, column=7).fill = color_cell
#                     ws.cell(row=row, column=7).font = text_color
#                     ws.cell(row=row, column=7).value = "-"
#
#                     ws.cell(row=row, column=8).alignment = align_cell
#                     ws.cell(row=row, column=8).border = border_cell
#                     ws.cell(row=row, column=8).fill = color_cell
#                     ws.cell(row=row, column=8).font = text_color
#                     ws.cell(row=row, column=8).value = round(d.quantity_niu, 2)
#
#                     ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=9).border = border_cell
#                     ws.cell(row=row, column=9).fill = color_cell
#                     ws.cell(row=row, column=9).font = text_color
#                     ws.cell(row=row, column=9).value = round(d.quantity_remaining, 2)
#
#                     ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=10).border = border_cell
#                     ws.cell(row=row, column=10).fill = color_cell
#                     ws.cell(row=row, column=10).font = text_color
#                     ws.cell(row=row, column=10).value = "-"
#
#                     ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=11).border = border_cell
#                     ws.cell(row=row, column=11).fill = color_cell
#                     ws.cell(row=row, column=11).font = text_color
#                     ws.cell(row=row, column=11).value = d.amount()
#
#                     ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=12).border = border_cell
#                     ws.cell(row=row, column=12).fill = color_cell
#                     ws.cell(row=row, column=12).font = text_color
#                     ws.cell(row=row, column=12).value = d.balance_remaining()
#
#                     ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=13).border = border_cell
#                     ws.cell(row=row, column=13).fill = color_cell
#                     ws.cell(row=row, column=13).font = text_color
#                     ws.cell(row=row, column=13).value = d.price
#                 elif d.operation == 'E':
#                     color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
#                     text_color = Font(name='Calibri', size=9)
#                     border_cell = Border(left=Side(border_style="thin"),
#                                          right=Side(border_style="thin"),
#                                          top=Side(border_style="thin"),
#                                          bottom=Side(border_style="thin"))
#                     align_cell = Alignment(horizontal="center")
#
#                     ws.cell(row=row, column=2).alignment = align_cell
#                     ws.cell(row=row, column=2).border = border_cell
#                     ws.cell(row=row, column=2).fill = color_cell
#                     ws.cell(row=row, column=2).font = text_color
#                     ws.cell(row=row, column=2).value = d.order.create_at
#
#                     ws.cell(row=row, column=3).alignment = align_cell
#                     ws.cell(row=row, column=3).border = border_cell
#                     ws.cell(row=row, column=3).fill = color_cell
#                     ws.cell(row=row, column=3).font = text_color
#                     ws.cell(row=row, column=3).value = str(d.order.get_type_display()).upper() + "Nº " + str(
#                         d.order.number)
#
#                     ws.cell(row=row, column=4).alignment = align_cell
#                     ws.cell(row=row, column=4).border = border_cell
#                     ws.cell(row=row, column=4).fill = color_cell
#                     ws.cell(row=row, column=4).font = text_color
#                     ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()
#
#                     ws.cell(row=row, column=5).alignment = align_cell
#                     ws.cell(row=row, column=5).border = border_cell
#                     ws.cell(row=row, column=5).fill = color_cell
#                     ws.cell(row=row, column=5).font = text_color
#                     ws.cell(row=row, column=5).value = str(d.quantity)
#
#                     ws.cell(row=row, column=6).alignment = align_cell
#                     ws.cell(row=row, column=6).border = border_cell
#                     ws.cell(row=row, column=6).fill = color_cell
#                     ws.cell(row=row, column=6).font = text_color
#                     ws.cell(row=row, column=6).value = str(d.get_unit_display())
#
#                     ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=7).border = border_cell
#                     ws.cell(row=row, column=7).fill = color_cell
#                     ws.cell(row=row, column=7).font = text_color
#                     ws.cell(row=row, column=7).value = d.quantity_niu
#
#                     ws.cell(row=row, column=8).alignment = align_cell
#                     ws.cell(row=row, column=8).border = border_cell
#                     ws.cell(row=row, column=8).fill = color_cell
#                     ws.cell(row=row, column=8).font = text_color
#                     ws.cell(row=row, column=8).value = "-"
#
#                     ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=9).border = border_cell
#                     ws.cell(row=row, column=9).fill = color_cell
#                     ws.cell(row=row, column=9).font = text_color
#                     ws.cell(row=row, column=9).value = d.quantity_remaining
#
#                     ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=10).border = border_cell
#                     ws.cell(row=row, column=10).fill = color_cell
#                     ws.cell(row=row, column=10).font = text_color
#                     ws.cell(row=row, column=10).value = d.amount()
#
#                     ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=11).border = border_cell
#                     ws.cell(row=row, column=11).fill = color_cell
#                     ws.cell(row=row, column=11).font = text_color
#                     ws.cell(row=row, column=11).value = "-"
#
#                     ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=12).border = border_cell
#                     ws.cell(row=row, column=12).fill = color_cell
#                     ws.cell(row=row, column=12).font = text_color
#                     ws.cell(row=row, column=12).value = d.balance_remaining()
#
#                     ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
#                     ws.cell(row=row, column=13).border = border_cell
#                     ws.cell(row=row, column=13).fill = color_cell
#                     ws.cell(row=row, column=13).font = text_color
#                     ws.cell(row=row, column=13).value = d.price
#             counter_for += 1
#             cont += 1
#             row += 1
#
#         # Establecer el nombre de mi archivo
#         nombre_archivo = "Kardex.xlsx"
#         # Definir el tipo de respuesta que se va a dar
#         response = HttpResponse(content_type="application/ms-excel")
#         contenido = "attachment; filename = {0}".format(nombre_archivo)
#         response["Content-Disposition"] = contenido
#         wb.save(response)
#         return response
#
def kardex_glp_excel(request, pk):
    user_id = request.user.id
    user_obj = User.objects.get(id=user_id)
    subsidiary_obj = get_subsidiary_by_user(user_obj)
